import urllib.request
import json
import cv2
import math
import numpy as np
import os
import random
import shutil
import matplotlib.image
import tensorflow as tf
from tensorflow import keras
import openpyxl
from JSONManager import JSONManager
from ModelUtils import ModelUtils
from Models.UNet import UNet
from Models.SegNet import SegNet
import distutils.util

class HyperparamStudy:

    def __init__(self, image_size=192, nn_type="UNet", n_layers=9, num_filters=[32,64,128,256,512], learning_rate=0.1, decay_rate=0.0, lr_decay_scheme="exp", lambd=0.0,
                 dropout_rate=[0.0,0.0], batch_size=64, batch_norm=[False,False], epochs=100, optimizer="adam", loss="binary_crossentropy", plot_flag=False):
        self.image_size = image_size
        self.nn_type = nn_type
        self.n_layers = n_layers
        self.num_filters = num_filters
        self.learning_rate = learning_rate
        self.decay_rate = decay_rate
        self.lr_decay_scheme = lr_decay_scheme
        self.lambd = lambd
        self.dropout_rate = dropout_rate
        self.batch_size = batch_size
        self.batch_norm = batch_norm
        self.saves_per_epoch = 2
        self.epochs = epochs
        self.optimizer = optimizer
        self.loss = loss
        self.plot_flag = plot_flag
        self.model = None
        self.training_score = 0
        self.validation_score = 0
        self.iou = 0

        # most of those variables won't change, if they need to change, change by hand
        self.train_path = "train/"
        self.valid_path = "validate/"
        self.image_path = "Images/"
        self.json_path ="export-2020-05-15T13_40_03.223Z.json"
        self.sets = ['main set 1']
        self.inputs_path = 'Inputs/'
        self.labels_path = 'Labels/'
        self.data_pct = [0.9, 0.05, 0.05]
        self.save_path = "Models/" + nn_type + "n" + str(n_layers) + "nf" + "-".join([str(f) for f in num_filters]) + "lr" + str(learning_rate) + "dordown" + str(dropout_rate[0]) \
                         + "dorup" + str(dropout_rate[1]) + "dr" + str(decay_rate) + "lamb" + str(lambd) + "bs" + str(batch_size) + "e" + str(epochs)
        self.save_path = self.save_path.replace(".","_") + ".h5" # replaces decimal points with underscores to prevent mess-ups in file name
        self.load_path = "Models/" + nn_type + "n" + str(n_layers) + "nf" + "-".join([str(f) for f in num_filters]) + "lr" + str(learning_rate) + "dordown" + str(dropout_rate[0]) \
                         + "dorup" + str(dropout_rate[1]) + "dr" + str(decay_rate) + "lamb" + str(lambd) + "bs" + str(batch_size) + "e" + str(epochs)
        self.load_path = self.load_path.replace(".","_") + ".h5" # replaces decimal points with underscores to prevent mess-ups in file name
        self.output_path = "Hyperparameters_Data.xlsx"


    def runner(self):
        JM = JSONManager(self.json_path, self.sets, self.inputs_path, self.labels_path, self.data_pct)
        MU = ModelUtils()

        x_train, y_train = JM.load_dataset(self.train_path+self.inputs_path, self.train_path+self.labels_path)
        x_val, y_val = JM.load_dataset(self.valid_path+self.inputs_path, self.valid_path+self.labels_path)
        x_train, y_train = JM.normalize_dataset(x_train,y_train)
        y_train2 = y_train[:, :, :, 0]
        y_train = np.expand_dims(y_train2, axis=-1)
        x_val, y_val = JM.normalize_dataset(x_val,y_val)
        y_val2 = y_val[:, :, :, 0]
        y_val = np.expand_dims(y_val2, axis=-1)


        if self.nn_type == "UNet":
            num_filters = [2**(4+i) for i in range(math.ceil(self.n_layers/2))]  # number of filters in each layer is a power of 2 starting at 16 up to bottleneck
            # if self.batch_norm:
            #     self.model = UNetBatchNorm(num_filters, self.image_size, self.lambd, self.dropout_rate)
            # else:
            self.model = UNet(num_filters, self.image_size, self.lambd, self.dropout_rate, self.batch_norm)
        elif self.nn_type == "SegNet":
            num_filters = [2**(4+i) for i in range(4)]
            self.model = SegNet(num_filters, self.image_size)
        #elif self.nn_type == "ResNet":
            #self.model = ResNet(, self.image_size)
        else:
            raise Exception("Invalid neural network architecture entered.")
        self.model = self.model.configure()

        # Learning rate decay options
        def exp_decay(epoch):
            learning_rate = self.learning_rate* np.exp(-self.decay_rate*epoch)
            return learning_rate

        def step_decay(epoch):
            epochs_drop = 10.0
            learning_rate = self.learning_rate * math.pow(self.decay_rate, math.floor((1+epoch)/epochs_drop))
            return learning_rate

        if self.decay_rate == 0.0:
            if self.optimizer.lower() == "adam":
                opt = keras.optimizers.Adam(learning_rate=self.learning_rate)
            elif self.optimizer.lower() == "sgd":
                opt = keras.optimizers.SGD(learning_rate=self.learning_rate)
            else:
                raise Exception("Please enter a valid optimizer.")
            self.model.compile(optimizer=opt, loss=self.loss, metrics=["acc",tf.keras.metrics.MeanIoU(num_classes=2)])
            self.model.summary()
            #Callbacks To Save every certain epochs
            filepath = "Models/" + self.nn_type + "n" + str(self.n_layers) + "nf" + "-".join([str(f) for f in num_filters]) + "lr" + str(self.learning_rate) + "dordown" + str(self.dropout_rate[0]) + "dorup" + str(self.dropout_rate[1]) + "dr" + str(self.decay_rate) + "lamb" + str(self.lambd) + "bs" + str(self.batch_size) + "e"
            filepath = filepath + "{epoch: 03d}"
            filepath = filepath.replace(".", "_") + ".h5"  # replaces decimal points with underscores to prevent mess-ups in file name
            cp = tf.keras.callbacks.ModelCheckpoint(filepath=filepath, period=int(round(self.epochs/self.saves_per_epoch)))
            self.model.fit(x=x_train, y=y_train, batch_size=self.batch_size, epochs=self.epochs, verbose=2)


        elif self.lr_decay_scheme == "time":
            if self.optimizer.lower() == "adam":
                opt = keras.optimizers.Adam(learning_rate=learning_rate, decay=self.decay_rate)
            elif self.optimizer.lower() == "sgd":
                opt = keras.optimizers.SGD(learning_rate=learning_rate, decay=self.decay_rate)
            else:
                raise Exception("Please enter a valid optimizer.")
            self.model.compile(optimizer=opt, loss=self.loss, metrics=["acc",tf.keras.metrics.MeanIoU(num_classes=2)])
            self.model.summary()
            # Callbacks To Save every certain epochs
            filepath = "Models/" + self.nn_type + "n" + str(self.n_layers) + "nf" + "-".join([str(f) for f in num_filters]) + "lr" + str(self.learning_rate) + "dordown" + str(self.dropout_rate[0]) + "dorup" + str(self.dropout_rate[1]) + "dr" + str(self.decay_rate) + "lamb" + str(self.lambd) + "bs" + str(self.batch_size) + "e"
            filepath = filepath + "{epoch: 03d}"
            filepath = filepath.replace(".","_") + ".h5"  # replaces decimal points with underscores to prevent mess-ups in file name
            cp = tf.keras.callbacks.ModelCheckpoint(filepath=filepath, period=int(round(self.epochs / self.saves_per_epoch)))
            self.model.fit(x=x_train, y=y_train, batch_size=self.batch_size, epochs=self.epochs, verbose=2)

        else:
            if self.lr_decay_scheme == "exp":
                callback = tf.keras.callbacks.LearningRateScheduler(exp_decay)  # makes callback with exponentially decaying learning rate
                if self.optimizer.lower() == "adam":
                    opt = keras.optimizers.Adam(learning_rate=self.learning_rate)
                elif self.optimizer.lower() == "sgd":
                    opt = keras.optimizers.SGD(learning_rate=self.learning_rate)
                else:
                    raise Exception("Please enter a valid optimizer.")
            elif self.lr_decay_scheme == "step":
                callback = tf.keras.callbacks.LearningRateScheduler(step_decay)  # makes callback with step-down decaying learning rate
                if self.optimizer.lower() == "adam":
                    opt = keras.optimizers.Adam(learning_rate=self.learning_rate)
                elif self.optimizer.lower() == "sgd":
                    opt = keras.optimizers.SGD(learning_rate=self.learning_rate)
                else:
                    raise Exception("Please enter a valid optimizer.")
            else:
                raise Exception("Please enter a valid learning rate decay scheme.")
            self.model.compile(optimizer=opt, loss=self.loss, metrics=["acc",tf.keras.metrics.MeanIoU(num_classes=2)])
            self.model.summary()
            # Callbacks To Save every certain epochs
            filepath = "Models/" + self.nn_type + "n" + str(self.n_layers) + "nf" + "-".join([str(f) for f in num_filters]) + "lr" + str(self.learning_rate) + "dordown" + str(self.dropout_rate[0]) + "dorup" + str(self.dropout_rate[1]) + "dr" + str(self.decay_rate) + "lamb" + str(self.lambd) + "bs" + str(self.batch_size) + "e"
            filepath = filepath + "{epoch: 03d}"
            filepath = filepath.replace(".","_") + ".h5"  # replaces decimal points with underscores to prevent mess-ups in file name
            cp = tf.keras.callbacks.ModelCheckpoint(filepath=filepath,period=int(round(self.epochs / self.saves_per_epoch)))
            self.model.fit(x=x_train, y=y_train, batch_size=self.batch_size,epochs=self.epochs, verbose=2)

        self.training_score = self.model.evaluate(x_train, y_train, verbose=0)[1]
        self.validation_score = self.model.evaluate(x_val, y_val, verbose=0)[1]
        self.iou = tf.keras.metrics.MeanIoU(num_classes=2).result().numpy()
        MU.save_model(self.model, self.save_path)

        if self.plot_flag:
            MU.show_validation(self.model,x_val, y_val)
            MU.show_train(self.model,x_train, y_train)
        # make file of 9 validation images (three 3x3 grids with .png, true labels, model labels)
        img = np.zeros((580,1760,3))  # 2 pixel wide red line separating imgs in grid, 10 pixel wide red line separating grids
        img[:,:,0] = 255  # make all red
        for i in range(3):
            for j in range(3):
                for k in range(3):
                    if i == 0:  # first grid plot model labels
                        result = self.model.predict(x_val[3*j+k:3*j+k+1, :, :, :])
                        _, result = cv2.threshold(result[0], 0.50, 255, cv2.THRESH_BINARY)
                        img[j*194:j*194+192,k*194+i*590:k*194+192+i*590,:] = np.repeat(result[:,:,np.newaxis],3,axis=2).astype('uint8')
                    elif i == 1:  # second grid plot true labels
                        img[j*194:j*194+192,k*194+i*590:k*194+192+i*590,:] = y_val[3*j+k].astype('uint8')
                    elif i == 2:  # third grid plot png
                        img[j*194:j*194+192,k*194+i*590:k*194+192+i*590,:] = x_val[3*j+k,:,:,:].astype('uint8')
        if not os.path.isdir("Images/"):
            os.mkdir("Images/")
        cv2.imwrite(self.image_path+self.save_path[7:-2]+"png", img)
        return


    def write(self):
        # open self.output_path, if empty file, add column titles, go to new line, add accuracies & other info for current model
        if not os.path.exists(self.output_path):
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Hyperparam Accuracies"
            ws1['A1'] = 'Neural Net Type'
            ws1['B1'] = 'Number of Layers'
            ws1['C1'] = 'Learning Rate'
            ws1['D1'] = 'Decay Rate'
            ws1['E1'] = 'Decay Scheme'
            ws1['F1'] = 'Lambda'
            ws1['G1'] = 'Dropout Rate'
            ws1['H1'] = 'Mini-batch Size'
            ws1['I1'] = 'Epochs'
            ws1['J1'] = 'Optimizer'
            ws1['K1'] = 'Loss'
            ws1['L1'] = 'Average Training Error'
            ws1['M1'] = 'Average Validation Error'
            ws1['N1'] = 'IoU'
            ws1['O1'] = 'Batch Norm Booleans'
            ws1['P1'] = 'Number of Filters'
        else:
            wb = openpyxl.load_workbook(filename=self.output_path)
            ws1 = wb.active
        row = ws1.max_row + 1  # find first empty row and add new data to it
        ws1["A"+str(row)] = self.nn_type
        ws1["B"+str(row)] = self.n_layers
        ws1["C"+str(row)] = self.learning_rate
        ws1["D"+str(row)] = self.decay_rate
        ws1["E"+str(row)] = self.lr_decay_scheme
        ws1["F"+str(row)] = self.lambd
        ws1["G"+str(row)] = ', '.join([str(dr) for dr in self.dropout_rate])
        ws1["H"+str(row)] = self.batch_size
        ws1["I"+str(row)] = self.epochs
        ws1["J"+str(row)] = self.optimizer
        ws1["K"+str(row)] = self.loss
        ws1["L"+str(row)] = self.training_score
        ws1["M"+str(row)] = self.validation_score
        ws1["N"+str(row)] = self.iou
        ws1["O"+str(row)] =  ', '.join([str(bn) for bn in self.batch_norm])
        ws1["P"+str(row)] =  ', '.join([str(num) for num in self.num_filters])
        wb.save(filename=self.output_path)
        return


if __name__ == '__main__':
    # Fill these np.arrays with as many options as user wishes (should all be same length)
    # image_size = [192]*9
    # nn_type = ["UNet"]*9  # can use syntax ["UNet"]*len(image_size) or ["UNet" for i in range(len(image_size))] to make a list of repeated nn architectures
    # n_layers = 9*np.ones((9,))
    # learning_rate = 0.001*np.ones((9,))
    # dropout_rate = np.array([[0.6,0.1],[0.6,0.1],[0.6,0.1],[0.7,0.1],[0.7,0.1],[0.7,0.1],[0.75,0.1],[0.75,0.1],[0.75,0.1]])
    # decay_rate = np.zeros((9,))
    # lr_decay_scheme = ["exp"]*9
    # lambd = np.array([0.0,0.0,0.1,0.0,0.0,0.1,0.0,0.0,0.1]) * 1e-4
    # batch_size = np.array([32,16,32,32,16,32,32,16,32])
    # batch_norm = [[False,False]]*9
    # epochs = [100] + [70]*8
    # optimizer = ["adam"]*9
    # loss = ["binary_crossentropy"]*9
    case_path = "Cases_To_Run.xlsx"
    wb = openpyxl.load_workbook(filename=case_path)
    ws1 = wb.active
    image_size = 192
    nn_type = []
    n_layers = []
    learning_rate = []
    dropout_rate = []
    decay_rate = []
    lr_decay_scheme = []
    lambd = []
    batch_size = []
    batch_norm = []
    epochs = []
    optimizer = []
    loss = []
    num_filters = []
    for row in range(2,ws1.max_row+1):
        nn_type.append(ws1['A'+str(row)].value)
        n_layers.append(ws1['B'+str(row)].value)
        learning_rate.append(ws1['C'+str(row)].value)
        decay_rate.append(ws1['D'+str(row)].value)
        lr_decay_scheme.append(ws1['E'+str(row)].value)
        lambd.append(ws1['F'+str(row)].value)
        dropout_rate.append([float(i) for i in ws1['G'+str(row)].value.split(",")])
        batch_size.append(ws1['H'+str(row)].value)
        epochs.append(ws1['I'+str(row)].value)
        optimizer.append(ws1['J'+str(row)].value)
        loss.append(ws1['K'+str(row)].value)
        batch_norm.append([bool(distutils.util.strtobool(i)) for i in ws1['L'+str(row)].value.split(",")])
        num_filters.append([float(i) for i in ws1['M'+str(row)].value.split(",")])
    plot_flag = False

    num_iter = ws1.max_row-1

    for i in range(num_iter):
        HPS = HyperparamStudy(image_size=image_size, nn_type=nn_type[i], n_layers=n_layers[i], num_filters=num_filters[i], learning_rate=learning_rate[i],
                              dropout_rate=dropout_rate[i], decay_rate=decay_rate[i], lr_decay_scheme=lr_decay_scheme[i],lambd=lambd[i],
                              batch_size=batch_size[i], batch_norm=batch_norm[i], epochs=epochs[i], optimizer=optimizer[i], loss=loss[i],plot_flag=plot_flag)
        HPS.runner()
        HPS.write()